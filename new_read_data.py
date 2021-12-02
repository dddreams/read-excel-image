from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl_image_loader import SheetImageLoader
import os
from PIL import Image

# 判断是否是文件和判断文件是否存在


def isfile_exist(file_path):
    if file_path is None:
        return False
    if not os.path.isfile(file_path):
        print("It's not a file or no such file exist ! %s" % file_path)
        return False
    else:
        return True


def get_size(file):
    # 获取文件大小:KB
    size = os.path.getsize(file)
    return size / 1024


def get_outfile(infile, outfile):
    if outfile:
        return outfile
    dir, suffix = os.path.splitext(infile)
    outfile = '{}{}'.format(dir, suffix)
    return outfile


def compress_image(infile, outfile='', mb=50, step=10, quality=60):
    """不改变图片尺寸压缩到指定大小
    :param infile: 压缩源文件
    :param outfile: 压缩文件保存地址
    :param mb: 压缩目标，KB
    :param step: 每次调整的压缩比率
    :param quality: 初始压缩比率
    :return: 压缩文件地址，压缩文件大小
    """
    o_size = get_size(infile)
    if o_size <= mb:
        return infile
    outfile = get_outfile(infile, outfile)
    while o_size > mb:
        im = Image.open(infile)
        im.save(outfile, quality=quality)
        if quality - step < 0:
            break
        quality -= step
        o_size = get_size(outfile)
    return outfile, get_size(outfile)


def resize_image(infile, outfile='', x_s=300):
    """修改图片尺寸
    :param infile: 图片源文件
    :param outfile: 重设尺寸文件保存地址
    :param x_s: 设置的宽度
    :return:
    """
    im = Image.open(infile)
    x, y = im.size
    y_s = int(y * x_s / x)
    out = im.resize((x_s, y_s), Image.ANTIALIAS)
    outfile = get_outfile(infile, outfile)
    out.save(outfile)


def read_files(file_path, image_target):
    #wb, ws, image_loader = None, None, None
    sql_arr, log_arr, error_arr = [], [], []

    wb = load_workbook(file_path)
    ws = wb.active
    image_loader = SheetImageLoader(ws)

    num = ws.max_row
    print(num)
    for i in range(2, num + 1):  # 2,3,4,5 ... 94
        name = str(ws['A'+str(i)].value).replace(' ', '').replace('\n','')
        sex = '1' if str(ws['B'+str(i)].value).replace(' ', '').replace('\n', '') == '男' else '0'
        phone = str(ws['C'+str(i)].value).replace(' ', '').replace('\n', '')

        photo = ''
        remark = ''
        img_size = 0
        image_path = ''

        if image_loader.image_in('D' + str(i)):
            image = image_loader.get('D' + str(i))
            image_path = image_target + phone + ".png"
            if not os.path.exists(image_target):
                os.makedirs(image_target)
            image = image.convert('RGB')
            image.save(image_path)

            photo = image_path
            img_size = get_size(image_path)

            # 照片压缩
            if img_size > 200:
                resize_image(image_path)
                compress_image(image_path)
            img_size = get_size(image_path)
        else:
            photo = ''
            remark += '照片格式不正确或未读取到,'

        if img_size > 200:
            remark += '照片大小大于200K,'
            photo = ''
        if len(phone) > 11:
            phone = phone[0:11]
        if phone == '':
            remark += '手机号码未采集,'
        if name == '':
            remark += '姓名未采集,'

        log_str = name + "|" + sex + "|" + phone + "|" + remark
        if remark != '':
            error_arr.append(log_str + '\n')
        log_arr.append('-------------------------------------------------------------------------------' + '\n')
        log_arr.append(log_str + '\n')
        log_arr.append(image_path + '\n')
        log_arr.append(file_path + '\n')

        sql_list = []
        sql_list.append(" insert into users(name, sex, phone, photo, remark)")
        sql_list.append(" values ('%s', '%s', '%s', '%s', '%s');")
        sql_str = str(''.join(sql_list)) % (name, sex, phone, photo, remark)
        sql_arr.append(sql_str + '\n')

    image_loader._images.clear()
    return sql_arr, log_arr, error_arr


if __name__ == '__main__':
    source_root = 'E:\\read-excel-image\\demo'
    target_root = 'E:\\read-excel-image\\target\\'

    sql_file = open(target_root + 'sql.sql', mode='w', encoding='utf-8')
    log_file = open(target_root + 'log.txt', mode='w', encoding='utf-8')
    error_file = open(target_root + 'error.txt', mode='w', encoding='utf-8')

    # workbook = xlwt.Workbook(encoding='ascii')
    # worksheet = workbook.add_sheet('存在问题的数据', cell_overwrite_ok=True)
    wb = Workbook()
    ws = wb.create_sheet("存在问题的数据", 0)

    index = 1
    for root, dirs, files in os.walk(source_root):
        for file in files:
            print(os.path.join(root, file))
            ddir = root.split(os.sep)[-1]
            sql_data, log_data, error_data = read_files(
                os.path.join(root, file), target_root + ddir + os.sep)

            sql_file.write(''.join(sql_data))
            log_file.write(''.join(log_data))
            error_file.write(''.join(error_data))

            for i in range(len(error_data)):
                index = index + 1
                arr_list = error_data[i].split("|")
                for j in range(len(arr_list)):
                    ws.cell(row = index, column= j+1, value = arr_list[j])

    sql_file.close()
    log_file.close()
    error_file.close()
    wb.save(target_root + '存在问题的数据.xlsx')
